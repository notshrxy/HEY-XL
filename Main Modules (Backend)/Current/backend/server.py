🧭 NaviBot – Smart Campus Navigation System
© 2025 Shreyas | Student of Sathyabama Institute of Science and Technology

import os
import uuid
from pathlib import Path
from typing import Dict, Optional, Any

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

# Ensure parent directory (Current/) is importable when running from backend/
import sys
CURRENT_DIR = Path(__file__).resolve().parent
PARENT_DIR = CURRENT_DIR.parent
if str(PARENT_DIR) not in sys.path:
    sys.path.insert(0, str(PARENT_DIR))

# --- Voice workflow imports (force lightweight SR-only path; no Whisper) ---
import numpy as np
import speech_recognition as sr
from module_voice_input import (
    record_audio,
    get_default_device_index,
    prompt_for_device_choice,
)
VOICE_SOURCE = "module_voice_input (no-whisper)"

from module_speaker_id import ensure_known_speaker
from module_parse_command import parse_command, set_speak_function
from module_excel_handler import ExcelHandler
import openpyxl
from tkinter import Tk, filedialog


app = Flask(__name__)
CORS(app)

UPLOAD_ROOT = Path(__file__).parent / "uploads"
UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)

sessions: Dict[str, Dict] = {}

# --- Global state for voice workflow ---
excel_instance: Optional[ExcelHandler] = None
sheet_selected: bool = False
device_index: Optional[int] = None
LATEST_FILE: Optional[Path] = None


def speak(text: str):
    print(text)


set_speak_function(speak)


def ensure_excel_loaded() -> ExcelHandler:
    global excel_instance, sheet_selected
    global LATEST_FILE
    if LATEST_FILE is None:
        raise RuntimeError("No Excel file uploaded. Use the Upload button before using the mic.")

    # If no instance or different file, (re)load
    if excel_instance is None or str(excel_instance.filename) != str(LATEST_FILE):
        excel_instance = ExcelHandler(str(LATEST_FILE))
        # Default to first sheet without any GUI prompt
        if excel_instance.wb.sheetnames:
            excel_instance.ws = excel_instance.wb[excel_instance.wb.sheetnames[0]]
        sheet_selected = True
    return excel_instance


def ensure_microphone_device() -> Optional[int]:
    global device_index
    if device_index is not None:
        return device_index
    try:
        choice = prompt_for_device_choice()
        device_index_local = choice["device_index"] if choice else None
        if device_index_local is None:
            device_index_local = get_default_device_index()
        device_index = device_index_local
    except Exception:
        device_index = get_default_device_index()
    return device_index


@app.get("/api/session/status")
def status():
    # Return shape compatible with older frontend code
    return jsonify({"sessions": len(sessions), "status": "running"}), 200


@app.post("/api/session/start")
def start_session():
    session_id = str(uuid.uuid4())
    sessions[session_id] = {"files": []}
    return jsonify({"sessionId": session_id}), 200


@app.post("/api/session/upload")
def upload():
    session_id = request.args.get("sessionId")
    if not session_id or session_id not in sessions:
        return jsonify({"error": "invalid sessionId"}), 400
    if "file" not in request.files:
        return jsonify({"error": "no file"}), 400

    f = request.files["file"]
    dest_dir = UPLOAD_ROOT / session_id
    dest_dir.mkdir(parents=True, exist_ok=True)
    # Always version uploads to avoid overwriting edited workbook
    stem = os.path.splitext(f.filename)[0]
    suffix = os.path.splitext(f.filename)[1] or ".xlsx"
    ts = uuid.uuid4().hex[:8]

    # If we already have a working file, back it up to history before switching
    prev_path = sessions[session_id].get("last_file")
    if prev_path and os.path.exists(prev_path):
        hist_dir = dest_dir / "history"
        hist_dir.mkdir(parents=True, exist_ok=True)
        prev_name = os.path.basename(prev_path)
        prev_stem, prev_suf = os.path.splitext(prev_name)
        backup_path = hist_dir / f"{prev_stem}_saved_{ts}{prev_suf or '.xlsx'}"
        try:
            import shutil
            shutil.copyfile(prev_path, backup_path)
        except Exception:
            pass

    # Save new upload with a unique name
    dest_path = dest_dir / f"{stem}_{ts}{suffix}"
    f.save(dest_path)

    sessions[session_id].setdefault("files", []).append(str(dest_path))
    sessions[session_id]["last_file"] = str(dest_path)
    global LATEST_FILE
    LATEST_FILE = dest_path
    return jsonify({"success": True, "path": str(dest_path)}), 200


@app.get("/api/session/open-local")
def open_local_file():
    """Let the user pick a local Excel file via OS dialog and use it in-place.
    Edits will be saved directly to this original path (no copies).
    """
    session_id = request.args.get("sessionId")
    if not session_id or session_id not in sessions:
        return jsonify({"error": "invalid sessionId"}), 400

    # Open file dialog
    try:
        root = Tk()
        root.withdraw()
        root.update()
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"),
                ("All files", "*.*"),
            ],
        )
        root.destroy()
    except Exception as e:
        return jsonify({"error": f"dialog_failed: {e}"}), 500

    if not path:
        return jsonify({"error": "cancelled"}), 400
    if not os.path.exists(path):
        return jsonify({"error": "not_found"}), 400

    # Point session to original file path
    sessions[session_id].setdefault("files", []).append(path)
    sessions[session_id]["last_file"] = path
    global LATEST_FILE
    LATEST_FILE = Path(path)
    return jsonify({"success": True, "path": path}), 200


@app.get("/api/session/preview")
def preview():
    session_id = request.args.get("sessionId")
    if not session_id or session_id not in sessions:
        return "Invalid session", 400
    last_file = sessions[session_id].get("last_file")
    if not last_file or not os.path.exists(last_file):
        files = sessions[session_id].get("files", [])
        items = "".join(f"<li>{os.path.basename(p)}</li>" for p in files)
        return (
            f"<html><body><h3>No file uploaded yet for this session.</h3><ul>{items or '<li>(none)</li>'}</ul></body></html>",
            200,
        )

    # Render first sheet as styled HTML table with auto-refresh
    wb = openpyxl.load_workbook(last_file, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows_html = []
    max_rows = min(sheet.max_row or 0, 200)
    max_cols = min(sheet.max_column or 0, 50)
    for r in range(1, max_rows + 1):
        cells = []
        for c in range(1, max_cols + 1):
            v = sheet.cell(row=r, column=c).value
            cells.append(f"<td>{'' if v is None else v}</td>")
        rows_html.append(f"<tr>{''.join(cells)}</tr>")

    html = f"""
    <html>
      <head>
        <title>Excel Live Preview</title>
        <meta http-equiv=\"refresh\" content=\"15\" />
        <style>
          body{{font-family:Segoe UI, Roboto, Arial, sans-serif;background:#f5f7fb;margin:0;padding:24px;}}
          .wrap{{max-width:1100px;margin:0 auto;background:#fff;border-radius:8px;box-shadow:0 4px 16px rgba(0,0,0,0.06);overflow:hidden;}}
          .header{{background:#213547;color:#fff;padding:14px 18px;display:flex;align-items:center;gap:10px;font-weight:600;}}
          .meta{{padding:10px 18px;color:#445; font-size:13px;border-bottom:1px solid #e8edf5;}}
          .btn{{display:inline-flex;align-items:center;gap:6px;background:#eef3ff;border:1px solid #c9d7ff;color:#275dff;padding:6px 10px;border-radius:6px;text-decoration:none;font-size:13px;}}
          .btn:hover{{background:#e5edff;}}
          .table-wrap{{padding:18px;}}
          table{{border-collapse:collapse;width:100%;}}
          td,th{{border:1px solid #e5e9f2;padding:10px;font-size:14px;}}
          thead td{{font-weight:700;background:#f0f4fa;}}
          tr:nth-child(even) td{{background:#fafbff;}}
          .right{{float:right}}
        </style>
      </head>
      <body>
        <div class=\"wrap\"> 
          <div class=\"header\">📊 Excel Live Preview</div>
          <div class=\"meta\">
            Session: {session_id} &nbsp; | &nbsp; File: {os.path.basename(last_file)}
            <a class=\"btn right\" href=\"?sessionId={session_id}\">🔄 Refresh</a>
          </div>
          <div class=\"table-wrap\">
            <table>
              {''.join(rows_html)}
            </table>
          </div>
        </div>
      </body>
    </html>
    """
    return html, 200


@app.get("/api/session/table")
def table_json():
    """Return the current sheet as JSON headers+rows for on-page table refresh."""
    session_id = request.args.get("sessionId")
    if not session_id or session_id not in sessions:
        return jsonify({"error": "invalid sessionId"}), 400
    last_file = sessions[session_id].get("last_file")
    if not last_file or not os.path.exists(last_file):
        return jsonify({"error": "no_file"}), 400

    wb = openpyxl.load_workbook(last_file, data_only=True)
    sheet = wb[wb.sheetnames[0]]

    max_rows = min(sheet.max_row or 0, 200)
    max_cols = min(sheet.max_column or 0, 50)

    headers = []
    if max_rows >= 1:
        for c in range(1, max_cols + 1):
            v = sheet.cell(row=1, column=c).value
            headers.append("" if v is None else v)

    rows = []
    for r in range(2, max_rows + 1):
        row_vals = []
        for c in range(1, max_cols + 1):
            v = sheet.cell(row=r, column=c).value
            row_vals.append("" if v is None else v)
        rows.append(row_vals)

    return jsonify({"headers": headers, "rows": rows}), 200


@app.get("/api/session/download")
def download_current_file():
    """Download the latest workbook for the given session as an attachment."""
    session_id = request.args.get("sessionId")
    if not session_id or session_id not in sessions:
        return jsonify({"error": "invalid sessionId"}), 400
    last_file = sessions[session_id].get("last_file")
    if not last_file or not os.path.exists(last_file):
        return jsonify({"error": "no_file"}), 400
    return send_file(
        last_file,
        as_attachment=True,
        download_name=os.path.basename(last_file),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# --- Voice endpoints on same server (port 8000) ---
@app.get("/api/voice/health")
def api_voice_health():
    return jsonify({"ok": True, "service": "voice", "version": 1, "source": VOICE_SOURCE}), 200


def capture_and_transcribe_in_memory(device: Optional[int], duration: float = 5.0, sample_rate: int = 16000) -> Optional[str]:
    """Record with module_voice_input.record_audio and transcribe via SpeechRecognition
    without creating temporary files (avoids WinError 32 on Windows)."""
    audio = record_audio(duration=duration, device_index=device, sample_rate=sample_rate)
    if audio is None:
        return None
    # Convert float32 [-1,1] to 16-bit PCM bytes
    pcm16 = (np.clip(audio, -1.0, 1.0) * 32767).astype(np.int16).tobytes()
    recognizer = sr.Recognizer()
    try:
        sr_audio = sr.AudioData(pcm16, sample_rate, 2)
        text = recognizer.recognize_google(sr_audio, language='en-US')
        return text.strip()
    except sr.UnknownValueError:
        return None
    except Exception:
        return None


@app.route("/api/voice/command", methods=["POST"])  # explicit to avoid confusion
def api_voice_command():
    result: Dict[str, Any] = {
        "status": "error",
        "steps": {
            "microphone": False,
            "speaker_identified": False,
            "listened": False,
            "parsed": False,
            "executed": False,
            "saved": False
        }
    }
    try:
        excel = ensure_excel_loaded()

        dev = ensure_microphone_device()
        result["steps"]["microphone"] = dev is not None

        user_name, score = ensure_known_speaker(
            duration=3.0,
            threshold=0.65,
            auto_enroll=True,
            speak_fn=speak,
            device=dev
        )
        result["speaker"] = {"name": user_name, "score": score, "recognized": user_name != "Unknown"}
        result["steps"]["speaker_identified"] = user_name != "Unknown"
        if user_name == "Unknown":
            result["message"] = "Voice not recognized or enrollment declined."
            return jsonify(result), 200

        speak("I'm listening. Please say your command.")
        # Simple SpeechRecognition-based capture using in-memory stream
        transcript = capture_and_transcribe_in_memory(device=dev)
        result["transcript"] = transcript
        result["steps"]["listened"] = transcript is not None
        if not transcript:
            result["message"] = "No speech detected or transcription failed."
            return jsonify(result), 200

        parsed = parse_command(transcript, excel)
        result["parsed"] = parsed
        result["steps"]["parsed"] = parsed is not None

        try:
            excel.wb.save(excel.filename)
            result["steps"]["saved"] = True
        except Exception as e:
            result["save_error"] = str(e)

        result["status"] = "ok"
        result["steps"]["executed"] = True
        result["message"] = "Command processed successfully"
        return jsonify(result), 200

    except Exception as e:
        result["error"] = str(e)
        return jsonify(result), 500


@app.post("/api/voice/text")
def api_text_command():
    """Execute a text command directly (no microphone or speaker ID)."""
    result: Dict[str, Any] = {
        "status": "error",
        "steps": {"parsed": False, "executed": False, "saved": False}
    }
    try:
        excel = ensure_excel_loaded()
        payload = request.get_json(silent=True) or {}
        text = (payload.get("text") or "").strip()
        if not text:
            result["message"] = "Empty command"
            return jsonify(result), 400

        parsed = parse_command(text, excel)
        result["parsed"] = parsed
        result["steps"]["parsed"] = parsed is not None

        try:
            excel.wb.save(excel.filename)
            result["steps"]["saved"] = True
        except Exception as e:
            result["save_error"] = str(e)

        result["status"] = "ok"
        result["steps"]["executed"] = True
        result["message"] = "Command processed successfully"
        return jsonify(result), 200
    except Exception as e:
        result["error"] = str(e)
        return jsonify(result), 500


def run(host: str = "127.0.0.1", port: int = 8000):
    app.run(host=host, port=port, debug=False)


if __name__ == "__main__":
    run()


