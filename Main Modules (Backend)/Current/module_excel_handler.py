🧭 NaviBot – Smart Campus Navigation System
© 2025 Shreyas | Student of Sathyabama Institute of Science and Technology

import openpyxl
import os
from openpyxl import Workbook
from tkinter import Tk, filedialog, messagebox, simpledialog
from fuzzywuzzy import fuzz
from pathlib import Path
import json
import zipfile
from openpyxl.utils.exceptions import InvalidFileException

from tkinter import Tk, filedialog
#Ask user for which file to pick
def ask_for_excel_file(app_folder=None):
    """Ask user to select or create an Excel file interactively."""
    root = Tk()
    root.withdraw()

    # Set initial directory to app folder if provided
    initial_dir = app_folder if app_folder else None

    file_path = filedialog.askopenfilename(
        title="Select an Excel file",
        initialdir=initial_dir,
        filetypes=[
            ("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"),
            ("Excel 2007+ files", "*.xlsx"),
            ("Excel Macro files", "*.xlsm"),
            ("Excel Template files", "*.xltx *.xltm"),
            ("All files", "*.*")
        ]
    )

    if not file_path:
        choice = messagebox.askyesno("No file selected", "No file selected. Do you want to create a new Excel file?")
        if choice:
            return create_new_excel_file(app_folder)
        else:
            messagebox.showinfo("Excel File Required", "We need at least one Excel file to proceed.")
            retry = messagebox.askyesno("Retry?", "Do you want to select an existing file again?")
            if retry:
                return ask_for_excel_file(app_folder)
            else:
                return None
    return file_path

def create_new_excel_file(app_folder=None):
    """Create a new Excel file after asking user for location + name."""
    from tkinter import filedialog
    initial_dir = app_folder if app_folder else None
    
    new_path = filedialog.asksaveasfilename(
        title="Create new Excel file",
        initialdir=initial_dir,
        defaultextension=".xlsx",
        filetypes=[
            ("Excel 2007+ files", "*.xlsx"),
            ("Excel Macro files", "*.xlsm"),
            ("Excel Template files", "*.xltx"),
            ("All files", "*.*")
        ]
    )
    if new_path:  # if user didn't cancel
        wb = Workbook()
        wb.save(new_path)
        print(f"New Excel file created at: {new_path}")
        return new_path
    return None

class ExcelHandler:
    def __init__(self, filename):
        self.filename = filename
        self.app_folder = self._ensure_app_folder()

        if os.path.exists(filename):
            # ✅ Open existing Excel file if it's a valid workbook; otherwise create a fresh one
            try:
                self.wb = openpyxl.load_workbook(filename)
                print(f"Opened existing file: {filename}")
            except (zipfile.BadZipFile, InvalidFileException):
                # File exists but is empty/corrupt or not a valid xlsx yet (e.g., NamedTemporaryFile)
                self.wb = openpyxl.Workbook()
                self.wb.save(filename)
                print(f"Reinitialized invalid Excel file: {filename}")
        else:
            # 🚀 Create new file
            self.wb = openpyxl.Workbook()
            self.wb.save(filename)
            print(f"Created new file: {filename}")

        # Handle sheets
        if not self.wb.sheetnames:
            # if no sheets, create one
            self.ws = self.wb.create_sheet("Sheet1")
        else:
            # default to first sheet, OR we'll override with choose_sheet()
            self.ws = self.wb[self.wb.sheetnames[0]]
        
        # Detect headers and setup data structures
        self.headers = self._detect_headers()
        self.student_data = self._load_student_data()

    #Ask the user which sheet to choose when multiple sheets exist
    def choose_sheet(self):
        """Ask user which sheet to use if multiple exist."""
        sheets = self.wb.sheetnames
        if len(sheets) == 1:
            return self.wb[sheets[0]]
        else:
            print("Available sheets:")
            for i, sheet in enumerate(sheets, start=1):
                print(f"{i}. {sheet}")
            choice = int(input("Enter sheet number to open: "))
            return self.wb[sheets[choice - 1]]
        
    # ---------------------------
    # Core Setup Functions
    # ---------------------------
    def get_or_create_workbook(self):
        try:
            workbook = openpyxl.load_workbook(self.filename)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active

        # Ask user what to name the sheet
            sheet_name = simpledialog.askstring("Input", "Enter name for the sheet:")
            if sheet_name:
                sheet.title = sheet_name

        # Ask user for column headers
            headers = []
            while True:
                header = simpledialog.askstring("Input", "Enter a column name (leave blank to finish):")
                if not header:
                    break
                headers.append(header)

            if headers:
                sheet.append(headers)

            workbook.save(self.filename)

        return workbook, sheet

    def save(self):
        """Save changes to workbook."""
        self.wb.save(self.filename)

    # ---------------------------
    # CRUD Functions
    # ---------------------------
    def add_score(self, student_name, subject, score):
        """Add a new score for a student and subject."""
        self.ws.append([student_name, subject, score])
        self.save()
        return f"✅ Added {score} for {student_name} in {subject}."

    def update_score(self, student_name, subject, new_score):
        """Update an existing score for a student and subject."""
        # Change 'self.sheet' to 'self.ws'
        for row in self.ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == student_name and row[1].value == subject:
                row[2].value = new_score
                # Fix: Use correct save method
                self.wb.save(self.filename)
                return f"🔄 Updated {student_name}'s {subject} score to {new_score}."
        return f"⚠️ No existing score found for {student_name} in {subject}."

    def delete_score(self, student_name, subject):
        """Delete a student’s score for a subject."""
        # Change 'self.sheet' to 'self.ws'
        for row in range(2, self.ws.max_row + 1):
            if (
                # Change 'self.sheet.cell' to 'self.ws.cell'
                self.ws.cell(row=row, column=1).value == student_name and
                self.ws.cell(row=row, column=2).value == subject
            ):
                # Change 'self.sheet.delete_rows' to 'self.ws.delete_rows'
                self.ws.delete_rows(row)
                # Change 'self.save()' to 'self.wb.save(self.filename)'
                self.save()
                return f"🗑️ Deleted {student_name}'s {subject} score."
        return f"⚠️ No score found for {student_name} in {subject}."

    def get_score(self, student_name, subject):
        """Retrieve a student’s score for a subject."""
        # Change 'self.sheet' to 'self.ws'
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row[0] == student_name and row[1] == subject:
                return f"📊 {student_name}'s score in {subject}: {row[2]}"
        return f"⚠️ No score found for {student_name} in {subject}."

    def get_all_scores(self, student_name):
        """Retrieve all subjects and scores for a student."""
        scores = []
        # Change 'self.sheet' to 'self.ws'
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row[0] == student_name:
                scores.append((row[1], row[2]))  # (subject, score)

        if not scores:
            return f"⚠️ No records found for {student_name}."

        result = f"📒 Scores for {student_name}:\n"
        for subject, score in scores:
            result += f" - {subject}: {score}\n"
        return result.strip()

    # ---------------------------
    # New Enhanced Methods
    # ---------------------------
    
    def _ensure_app_folder(self):
        """Create dedicated app folder if it doesn't exist"""
        app_folder = os.path.join(os.path.expanduser("~"), "ExcelVoiceApp")
        os.makedirs(app_folder, exist_ok=True)
        print(f"📁 App folder: {app_folder}")
        return app_folder

    def _detect_headers(self):
        """Detect column headers from the first row"""
        if self.ws.max_row > 0:
            headers = []
            for cell in self.ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    break
            return headers
        return []

    def _load_student_data(self):
        """Load student data for fuzzy matching"""
        student_data = {}
        if self.ws.max_row > 1:  # Skip header row
            for row_num, row in enumerate(self.ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0]:  # If first column has data
                    student_name = str(row[0]).strip()
                    student_data[student_name.lower()] = {
                        'name': student_name,
                        'row': row_num
                    }
        return student_data

    def find_student_row(self, student_name, threshold=80):
        """Find row number for a student using fuzzy matching"""
        # Refresh cached data in case the sheet was modified externally
        self.student_data = self._load_student_data()
        student_name = student_name.strip()
        
        # First try exact match
        for name, data in self.student_data.items():
            if name.lower() == student_name.lower():
                return data['row']
        
        # Then try fuzzy matching
        best_match = None
        best_score = 0
        
        for name, data in self.student_data.items():
            score = fuzz.ratio(student_name.lower(), name.lower())
            if score > best_score and score >= threshold:
                best_match = data
                best_score = score
        
        if best_match:
            print(f"🔍 Found student: '{best_match['name']}' (match: {best_score}%)")
            return best_match['row']
        
        # If no match found, return None
        return None

    def find_subject_column(self, subject_name, threshold=80):
        """Find column letter for a subject using fuzzy matching"""
        # Refresh headers in case the sheet was modified externally
        self.headers = self._detect_headers()
        subject_name = subject_name.strip().upper()
        
        # First try exact match
        for i, header in enumerate(self.headers, 1):
            if header.upper() == subject_name:
                return openpyxl.utils.get_column_letter(i)
        
        # Then try fuzzy matching
        best_match = None
        best_score = 0
        
        for i, header in enumerate(self.headers, 1):
            score = fuzz.ratio(subject_name, header.upper())
            if score > best_score and score >= threshold:
                best_match = (header, i)
                best_score = score
        
        if best_match:
            print(f"🔍 Found subject: '{best_match[0]}' (match: {best_score}%)")
            return openpyxl.utils.get_column_letter(best_match[1])
        
        return None

    def update_cell_value(self, student_name, subject, value):
        """Update specific cell instead of appending rows"""
        # Ensure latest headers and student map
        self.headers = self._detect_headers()
        self.student_data = self._load_student_data()
        # Find student row
        student_row = self.find_student_row(student_name)
        if not student_row:
            return f"❌ Student '{student_name}' not found. Please check the name or add them first."
        
        # Find subject column
        subject_col = self.find_subject_column(subject)
        if not subject_col:
            return f"❌ Subject '{subject}' not found. Available subjects: {', '.join(self.headers)}"
        
        # Update the cell
        cell = self.ws[f"{subject_col}{student_row}"]
        old_value = cell.value
        cell.value = value
        
        # Save the file
        self.wb.save(self.filename)
        
        return f"✅ Updated {student_name}'s {subject} from '{old_value}' to '{value}' in cell {subject_col}{student_row}"

    def add_student_if_not_exists(self, student_name):
        """Add a new student row if they don't exist"""
        if self.find_student_row(student_name):
            return False  # Student already exists
        
        # Add new row with student name
        new_row = [student_name] + [""] * (len(self.headers) - 1)
        self.ws.append(new_row)
        
        # Update student data
        self.student_data[student_name.lower()] = {
            'name': student_name,
            'row': self.ws.max_row
        }
        
        self.wb.save(self.filename)
        return True

    def validate_subject(self, subject_name, threshold=80):
        """Check if subject exists in headers"""
        # Refresh headers in case they were populated after handler initialization
        self.headers = self._detect_headers()
        if not self.headers:
            return False, "No headers found in the Excel file"
        
        # Try exact match first
        for header in self.headers:
            if header.upper() == subject_name.upper():
                return True, f"Subject '{subject_name}' found"
        
        # Try fuzzy match
        best_match = None
        best_score = 0
        
        for header in self.headers:
            score = fuzz.ratio(subject_name.upper(), header.upper())
            if score > best_score and score >= threshold:
                best_match = header
                best_score = score
        
        if best_match:
            return True, f"Subject '{subject_name}' matched with '{best_match}' ({best_score}%)"
        
        return False, f"Subject '{subject_name}' not found. Available subjects: {', '.join(self.headers)}"
