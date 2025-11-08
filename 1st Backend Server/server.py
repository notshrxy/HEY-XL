#!/usr/bin/env python3
"""
Live Preview Server for Excel Voice App
Handles file uploads and serves live preview of Excel files
"""

import os
import sys
import json
import time
import signal
import threading
from pathlib import Path
from typing import Dict, Optional
import tempfile
import shutil

from flask import Flask, request, jsonify, send_file, render_template_string
from flask_cors import CORS
import openpyxl
from openpyxl.utils import get_column_letter

# Global variables for session management
sessions: Dict[str, Dict] = {}
app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# HTML template for the live preview
PREVIEW_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Live Preview</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            overflow: hidden;
        }
        .header {
            background: #2c3e50;
            color: white;
            padding: 20px;
            text-align: center;
        }
        .header h1 {
            margin: 0;
            font-size: 24px;
        }
        .session-info {
            background: #ecf0f1;
            padding: 10px 20px;
            border-bottom: 1px solid #bdc3c7;
            font-size: 14px;
            color: #7f8c8d;
        }
        .controls {
            padding: 20px;
            background: #f8f9fa;
            border-bottom: 1px solid #dee2e6;
        }
        .refresh-btn {
            background: #3498db;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 5px;
            cursor: pointer;
            font-size: 14px;
            transition: background 0.3s;
        }
        .refresh-btn:hover {
            background: #2980b9;
        }
        .refresh-btn:disabled {
            background: #bdc3c7;
            cursor: not-allowed;
        }
        .status {
            margin-left: 10px;
            font-size: 14px;
        }
        .status.success { color: #27ae60; }
        .status.error { color: #e74c3c; }
        .status.loading { color: #f39c12; }
        .content {
            padding: 20px;
            overflow-x: auto;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 0;
        }
        th, td {
            border: 1px solid #ddd;
            padding: 12px;
            text-align: left;
        }
        th {
            background-color: #34495e;
            color: white;
            font-weight: bold;
        }
        tr:nth-child(even) {
            background-color: #f2f2f2;
        }
        tr:hover {
            background-color: #e8f4f8;
        }
        .no-data {
            text-align: center;
            padding: 40px;
            color: #7f8c8d;
            font-style: italic;
        }
        .error-message {
            background: #e74c3c;
            color: white;
            padding: 15px;
            border-radius: 5px;
            margin: 20px 0;
        }
        .last-updated {
            text-align: right;
            font-size: 12px;
            color: #7f8c8d;
            margin-top: 20px;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 Excel Live Preview</h1>
        </div>
        <div class="session-info">
            Session ID: {{ session_id }} | File: {{ filename }}
        </div>
        <div class="controls">
            <button class="refresh-btn" onclick="refreshData()">🔄 Refresh</button>
            <span class="status" id="status"></span>
        </div>
        <div class="content" id="content">
            {{ content | safe }}
        </div>
        <div class="last-updated" id="lastUpdated">
            Last updated: {{ last_updated }}
        </div>
    </div>

    <script>
        let refreshInterval;
        
        function refreshData() {
            const statusEl = document.getElementById('status');
            const refreshBtn = document.querySelector('.refresh-btn');
            
            statusEl.textContent = 'Refreshing...';
            statusEl.className = 'status loading';
            refreshBtn.disabled = true;
            
            fetch('/api/session/refresh?sessionId={{ session_id }}')
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        document.getElementById('content').innerHTML = data.html;
                        document.getElementById('lastUpdated').textContent = 'Last updated: ' + data.last_updated;
                        statusEl.textContent = 'Refreshed successfully';
                        statusEl.className = 'status success';
                    } else {
                        statusEl.textContent = 'Refresh failed: ' + data.error;
                        statusEl.className = 'status error';
                    }
                })
                .catch(error => {
                    statusEl.textContent = 'Refresh failed: ' + error.message;
                    statusEl.className = 'status error';
                })
                .finally(() => {
                    refreshBtn.disabled = false;
                    setTimeout(() => {
                        statusEl.textContent = '';
                        statusEl.className = 'status';
                    }, 3000);
                });
        }
        
        // Auto-refresh every 5 seconds
        function startAutoRefresh() {
            refreshInterval = setInterval(refreshData, 5000);
        }
        
        // Stop auto-refresh when page is hidden
        document.addEventListener('visibilitychange', function() {
            if (document.hidden) {
                clearInterval(refreshInterval);
            } else {
                startAutoRefresh();
            }
        });
        
        // Start auto-refresh
        startAutoRefresh();
    </script>
</body>
</html>
"""

def excel_to_html(file_path: str) -> str:
    """Convert Excel file to HTML table."""
    try:
        if not os.path.exists(file_path):
            return '<div class="error-message">File not found</div>'
        
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        if ws.max_row == 0:
            return '<div class="no-data">No data in Excel file</div>'
        
        # Build HTML table
        html = '<table>\n'
        
        # Add header row
        if ws.max_row > 0:
            html += '  <tr>\n'
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                html += f'    <th>{cell_value if cell_value is not None else ""}</th>\n'
            html += '  </tr>\n'
        
        # Add data rows
        for row in range(2, ws.max_row + 1):
            html += '  <tr>\n'
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                html += f'    <td>{cell_value if cell_value is not None else ""}</td>\n'
            html += '  </tr>\n'
        
        html += '</table>'
        return html
        
    except Exception as e:
        return f'<div class="error-message">Error reading Excel file: {str(e)}</div>'

@app.route('/api/session/status')
def session_status():
    """Check if server is running."""
    return jsonify({"status": "running", "sessions": len(sessions)})

@app.route('/api/session/start', methods=['POST'])
def start_session():
    """Start a new preview session."""
    session_id = f"session_{int(time.time())}"
    sessions[session_id] = {
        'created_at': time.time(),
        'file_path': None,
        'last_updated': None
    }
    return jsonify({"sessionId": session_id})

@app.route('/api/session/upload', methods=['POST'])
def upload_file():
    """Upload Excel file to session."""
    session_id = request.args.get('sessionId')
    if not session_id or session_id not in sessions:
        return jsonify({"error": "Invalid session ID"}), 400
    
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400
    
    # Save uploaded file
    temp_dir = Path(tempfile.gettempdir()) / "excel_preview"
    temp_dir.mkdir(exist_ok=True)
    file_path = temp_dir / f"{session_id}_{file.filename}"
    file.save(file_path)
    
    # Update session
    sessions[session_id]['file_path'] = str(file_path)
    sessions[session_id]['last_updated'] = time.time()
    
    return jsonify({"success": True, "message": "File uploaded successfully"})

@app.route('/api/session/refresh')
def refresh_session():
    """Refresh session data."""
    session_id = request.args.get('sessionId')
    if not session_id or session_id not in sessions:
        return jsonify({"error": "Invalid session ID"}), 400
    
    session = sessions[session_id]
    if not session['file_path'] or not os.path.exists(session['file_path']):
        return jsonify({"error": "No file available"}), 400
    
    try:
        html = excel_to_html(session['file_path'])
        session['last_updated'] = time.time()
        last_updated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(session['last_updated']))
        
        return jsonify({
            "success": True,
            "html": html,
            "last_updated": last_updated
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/session/preview')
def preview_session():
    """Serve the preview page."""
    session_id = request.args.get('sessionId')
    if not session_id or session_id not in sessions:
        return "Invalid session ID", 400
    
    session = sessions[session_id]
    filename = "No file" if not session['file_path'] else os.path.basename(session['file_path'])
    
    # Get current content
    if session['file_path'] and os.path.exists(session['file_path']):
        content = excel_to_html(session['file_path'])
    else:
        content = '<div class="no-data">No file uploaded yet</div>'
    
    last_updated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(session.get('last_updated', time.time())))
    
    return render_template_string(
        PREVIEW_TEMPLATE,
        session_id=session_id,
        filename=filename,
        content=content,
        last_updated=last_updated
    )

@app.route('/api/session/cleanup', methods=['POST'])
def cleanup_session():
    """Clean up session and files."""
    session_id = request.args.get('sessionId')
    if session_id and session_id in sessions:
        session = sessions[session_id]
        if session['file_path'] and os.path.exists(session['file_path']):
            try:
                os.remove(session['file_path'])
            except:
                pass
        del sessions[session_id]
        return jsonify({"success": True})
    return jsonify({"error": "Session not found"}), 404

def cleanup_all_sessions():
    """Clean up all sessions and temporary files."""
    temp_dir = Path(tempfile.gettempdir()) / "excel_preview"
    if temp_dir.exists():
        try:
            shutil.rmtree(temp_dir)
        except:
            pass
    sessions.clear()

def signal_handler(signum, frame):
    """Handle shutdown signals."""
    print("\nShutting down Live Preview Server...")
    cleanup_all_sessions()
    sys.exit(0)

if __name__ == "__main__":
    # Register signal handlers for graceful shutdown
    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)
    
    # Clean up on startup
    cleanup_all_sessions()
    
    print("Starting Live Preview Server...")
    print("Server will be available at: http://127.0.0.1:8000")
    print("Press Ctrl+C to stop the server")
    
    try:
        app.run(host='127.0.0.1', port=8000, debug=False, threaded=True)
    except KeyboardInterrupt:
        print("\nServer stopped by user")
    finally:
        cleanup_all_sessions()
        print("Server cleanup completed")